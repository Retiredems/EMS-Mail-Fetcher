"""Export contacts to Excel (.xlsx) using openpyxl."""

from pathlib import Path
from typing import Sequence

from db.models import Contact


def export_contacts_excel(contacts: Sequence[Contact], output_path: Path) -> int:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl is not installed — run: pip install openpyxl")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    headers = ["Email", "Display Name", "Occurrences", "First Seen", "Last Seen", "Account ID"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D6A9F")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, c in enumerate(contacts, start=2):
        ws.cell(row=row_idx, column=1, value=c.email_address)
        ws.cell(row=row_idx, column=2, value=c.display_name or "")
        ws.cell(row=row_idx, column=3, value=c.occurrence_count)
        ws.cell(row=row_idx, column=4, value=c.first_seen_at.isoformat() if c.first_seen_at else "")
        ws.cell(row=row_idx, column=5, value=c.last_seen_at.isoformat() if c.last_seen_at else "")
        ws.cell(row=row_idx, column=6, value=c.account_id)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(output_path)
    return len(contacts)
